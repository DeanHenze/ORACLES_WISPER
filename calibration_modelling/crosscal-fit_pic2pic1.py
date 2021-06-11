# -*- coding: utf-8 -*-
"""
Created on Thu Mar 11 10:47:32 2021

@author: Dean

Cross calibration of Pic2 (Gulper) to Pic1 (Mako). This is done for all 2017 
and 2018 flights - where there is data - other than Aug. 12th and 13th, 2017. 
For those two flights, Mako's calibration was clearly off.

Notes:
    I get the statsmodels message:
        "The condition number is large, 8.17e+03. This might indicate that there are
        strong multicollinearity or other numerical problems." 
"""


# Built in:
import os, sys
import itertools

# Third party:
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import statsmodels.api as sm
from openpyxl import load_workbook

# My modules:
import ancillary as anc
if r'../..' not in sys.path: sys.path.insert(0,r'../..')
from my_python_modules import math_henze as hmath


"""
Load all WISPER (q,dD,d18O) data for either 2017 or 2018 with good data. 
Return as a pandas df.
"""
def get_wisperdata(year):

    path_data_dir = r"../WISPER_data/pic1_cal/"

    if year=='2017':
        dates_good = ['20170815','20170817','20170818','20170821','20170824',
                      '20170826','20170828','20170830','20170831','20170902']
    elif year=='2018':
        dates_good = ['20180927','20180930','20181003','20181007','20181010',
                      '20181012','20181015','20181017','20181019','20181021',
                      '20181023']
            
    fnames = ['WISPER_'+d+'_pic1-cal.ict' for d in dates_good]
    paths_dates = [path_data_dir+f for f in fnames] # Paths to each data file.
    
## Loop through each date and append data to a single pandas df:
    columns = ['h2o_tot1','h2o_tot2','dD_tot1',
               'dD_tot2','d18O_tot1','d18O_tot2'] # Keep these data columns.
    wisper = pd.DataFrame({}, columns=columns) # Append all data here.
    for p in paths_dates:
        data_temp = pd.read_csv(p) # Load.
        data_temp.replace(-9999, np.nan, inplace=True)
        #data_temp = data_temp.rolling(window=10).mean() # Running mean.
        #wisper = wisper.append(data_temp[columns], ignore_index=True)
        data_blocked = data_temp.groupby(lambda x: np.round(x/8)).mean()
        wisper = wisper.append(data_blocked[columns], ignore_index=True)
    
    return wisper.dropna(how='any') # Drop missing values


"""
Fit a line to Pic1 vs Pic2 vapor concentration. Line is constrained to pass
through the origin. 'df' is the pandas dataframe of wisper measurements.
"""
def q_xcal(df):
    return sm.OLS(df['h2o_tot1'], df['h2o_tot2'], missing='drop').fit()
    
    
"""
Produces powers of predictor vars for a candidate model to use with 
statsmodels. 

predictvars (pandas dataframe): Contains predictor vars for model.
pwr_range (dict): keys are a subset of the columns in predictvars. Elements are 
    2-tuples for min, max powers for that key. Powers can be negative.
"""
def get_poly_terms(predictvars, pwr_range):

    modelvars = pd.DataFrame({}) # Will hold all vars to use with model.
    
    for key in pwr_range.keys():
        # Powers (excluding 0) of predictor var to try:
        powers = list(range(pwr_range[key][0],pwr_range[key][1]+1))
        if 0 in powers: del powers[powers.index(0)]
        
        for i in powers: # Compute and append powers:
            modelvars[key+'^%i' % i] = predictvars[key]**i
            
    # Add constant offset var:
    modelvars['const'] = np.ones(len(predictvars))
            
    return modelvars
    

"""
Get cross-calibration formula for either dD or d18O.

Fit Pic1 isotope measurements to a polynomial of Pic2 log(q) and isotope 
measurements, including an interaction term. Use a weighted linear regression 
from the statsmodels package (weights are the water concentrations). 

'df' is the pandas dataframe of wisper measurements. 'iso' is the isotope to 
get cross-cal formula for (can be either 'dD' or 'd18O'). 'nord' is 3-tuple of 
ints, corresponding to the highest power to include for each predictor 
variable: logq, isotope-ratio, and their crossterm.
"""
def isoxcal_modelfit(df, iso, nord):

    # Pandas df of predictor vars:
    interterm = np.log(df['h2o_tot2'])*df[iso+'_tot2'] # Interaction term.
    predictvars = pd.DataFrame({'logq':np.log(df['h2o_tot2']),
                                iso:df[iso+'_tot2'],
                                'logq*'+iso:interterm})
    
    # Get all desired powers of each predictor var; add as new df columns:
    pwr_range = {'logq':(0,nord[0]), iso:(0,nord[1]), 
                 'logq*'+iso:(0,nord[2])}
    predictvars_poly = get_poly_terms(predictvars, pwr_range)
    
    # Return model fit:
    return sm.WLS(df[iso+'_tot1'], predictvars_poly, missing='drop', 
                      weights=df['h2o_tot2']).fit()

    
"""
The isotope cross-calibration model. Takes in model parameters and 
data for the predictor variables (logq, iso, logq*iso), and returns calibrated 
isotope ratios. iso is either 'dD' or 'd18O'.

predvars: dict-like containing values of the predictor vars (arrays of same length).
pars: fit parameters (dict-like). The keys for the fit parameters need to be 
    of the form 'var^n' where n is the power of the term in the model and 'var' 
    is the key for the appropriate variable in 'data'.
"""
def model_isoxcal(predvars, pars):
    terms = []
    
    for k in pars.keys():
        if k=='const':
            terms.append(pars[k]*np.ones(np.shape(predvars[list(predvars.keys())[0]])))
        else:
            pvar, power = k.split('^') # Predictor var name and power it's raised to.
            terms.append(pars[k]*predvars[pvar]**int(power))
    
    return np.sum(terms, axis=0)


def model_residual_map(iso, wisperdata, pars, logq_grid, iso_grid, ffact=1):
    # Get model predictions:
    logq = np.log(wisperdata['h2o_tot2'].values)
    predictorvars = {'logq':logq, 
                     iso:wisperdata[iso+'_tot2'].values, 
                     'logq*'+iso:logq*wisperdata[iso+'_tot2'].values, 
                     }
    modelresults = model_isoxcal(predictorvars, pars)
    # Model residuals:
    res = abs(modelresults - wisperdata[iso+'_tot1'])
    # Get RMSE 2d map using oversampling:
    return hmath.oversampler(res, logq, wisperdata[iso+'_tot2'], 
                             logq_grid, iso_grid, ffact=ffact)          
        
    
def draw_fitfig(year, wisperdata, pars_dD, pars_d18O):

    fig = plt.figure(figsize=(6.5,2.5))
    ax_D = plt.axes([0.125,0.2,0.29,0.75])
    cbax_D = plt.axes([0.435,0.2,0.02,0.625])
    ax_18O = plt.axes([0.62,0.2,0.29,0.75])
    cbax_18O = plt.axes([0.93,0.2,0.02,0.625])
    
    ## Some year-dependent contour/colormapping values:
    if year=='2017':
        # Min/max values for colormap renormalization:
        vmin_D = -650; vmax_D = 0; vmin_18O = -80; vmax_18O = 0
        # Contour levels for plotting model output:
        clevs_D = np.arange(-600,0,50); clevs_18O = np.arange(-80,0,5)
    if year=='2018':
        vmin_D = -200; vmax_D = -40; vmin_18O = -30; vmax_18O = -8
        clevs_D = np.arange(-200,0,15); clevs_18O = np.arange(-30,0,2)

    ## WISPER data scatter plots:
    ##-------------------------------------------------------------------------
        # Thin out the wisper data for better visuals:
    wisperthin = wisperdata.iloc[np.arange(0,len(wisperdata),10)]
        
    s_D = ax_D.scatter(np.log(wisperthin['h2o_tot2']), wisperthin['dD_tot2'], 
                       c=wisperthin['dD_tot1'], vmin=vmin_D, vmax=vmax_D, 
                       s=5)
    s_18O = ax_18O.scatter(np.log(wisperthin['h2o_tot2']), wisperthin['d18O_tot2'], 
                           c=wisperthin['d18O_tot1'], 
                           vmin=vmin_18O, vmax=vmax_18O, s=5)
    ##-------------------------------------------------------------------------

    ## Compute model-fit values and plot as contours:
    ##-------------------------------------------------------------------------
        # Predictor variable values to pass into model:
    if year=='2017':
        logq = np.linspace(np.log(100), np.log(30000), 200)
        logq_grid, dD_grid = np.meshgrid(logq, np.linspace(-450, -30, 100))
        logq_grid, d18O_grid = np.meshgrid(logq, np.linspace(-55, 0, 100))
    if year=='2018':
        logq = np.linspace(np.log(2000), np.log(30000), 200)
        logq_grid, dD_grid = np.meshgrid(logq, np.linspace(-200, -30, 100))
        logq_grid, d18O_grid = np.meshgrid(logq, np.linspace(-55, -30, 100))
    predictorvars = {'logq':logq_grid, 
                     'dD':dD_grid, 
                     'd18O':d18O_grid, 
                     'logq*dD':logq_grid*dD_grid, 
                     'logq*d18O':logq_grid*d18O_grid
                     }

        # Run model:
    modeldata_dD = model_isoxcal(predictorvars, pars_dD)
    modeldata_d18O = model_isoxcal(predictorvars, pars_d18O)
    
        # Contour plots of model output:
    ax_D.contour(logq_grid, dD_grid, modeldata_dD, 
                 levels=clevs_D, vmin=vmin_D, vmax=vmax_D, linewidths=2.5)
    ax_18O.contour(logq_grid, d18O_grid, modeldata_d18O, 
                   levels=clevs_18O, vmin=vmin_18O, vmax=vmax_18O, linewidths=2.5)
    ##-------------------------------------------------------------------------
    
    ## Include contours of model residuals
    ##-------------------------------------------------------------------------
        # Compute root-mean-square deviations:
    if year=='2017':
        reslevs_D = [2,5,15,30]; reslevs_18O = [0.2,0.5,1,2,4]; ffact=1.5
    if year=='2018': 
        reslevs_D = [1,2,5,10,20]; reslevs_18O = [0.2,0.5,1,2]; ffact=4
    
    res_dD = model_residual_map('dD', wisperdata, pars_dD, logq, dD_grid[:,0], ffact=ffact)  
    res_d18O = model_residual_map('d18O', wisperdata, pars_d18O, logq, d18O_grid[:,0], ffact=ffact)  
    
        # Contours:
    rescont_D = ax_D.contour(logq_grid, dD_grid, res_dD, 
                             levels=reslevs_D, colors='black', linewidths=1)
    rescont_18O = ax_18O.contour(logq_grid, d18O_grid, res_d18O, 
                                 levels=reslevs_18O, colors='black', linewidths=1)
    plt.clabel(rescont_D, inline=True, fmt='%i')
    plt.clabel(rescont_18O, inline=True, fmt='%0.1f')
    ##-------------------------------------------------------------------------

    ## Figure axes labels, limits, colobars, ...
    ##-------------------------------------------------------------------------
        # Results axes mods:
    if year=='2017':
        ax_D.set_xlim(4.5, 10.5); ax_D.set_ylim(-400, -20)
        ax_18O.set_xlim(4.5, 10.5); ax_18O.set_ylim(-50, 0)
    if year=='2018':
        ax_D.set_xlim(7.5, 10.5); ax_D.set_ylim(-180, -40)
        ax_18O.set_xlim(7.5, 10.5); ax_18O.set_ylim(-50, -32)

    ax_D.set_xlabel(r'log($q_2$[ppmv])', fontsize=12)
    ax_D.set_ylabel(r'$\delta D_2$'+u'(\u2030)', fontsize=12, labelpad=0)
    ax_18O.set_xlabel(r'log($q_2$[ppmv])', fontsize=12)
    ax_18O.set_ylabel(r'$\delta^{18} O_2$'+u'(\u2030)', fontsize=12, labelpad=0)   
    
        # Colorbar axes mods:
    fig.colorbar(s_D, cax=cbax_D)
    fig.colorbar(s_18O, cax=cbax_18O)
    
    cbax_D.set_title(r'$\delta D_1$'+'\n'+u'(\u2030)', fontsize=10)
    plt.setp(cbax_D.yaxis.get_majorticklabels(), 
             ha="center", va="center", rotation=-90, rotation_mode="anchor")
    
    cbax_18O.set_title(r'$\delta^{18} O_1$'+'\n'+u'(\u2030)', fontsize=10)
    plt.setp(cbax_18O.yaxis.get_majorticklabels(), 
         ha="center", va="center", rotation=-90, rotation_mode="anchor")
    
    ##-------------------------------------------------------------------------



## Save fit results to my calibration fits excel file:
##-----------------------------------------------------------------------------    

def get_fits():

    """
    Get cross-calibration formula fit parameters for water concentration and both 
    isotopologues for an ORACLES year (either '2017' or '2018'). Return as a 
    dictionary of pandas Series objects.
    """
    def get_fits_year(year):
            
        wisperdata = get_wisperdata(year)

        ## Fitting humidity is straightforward:
        model_q = q_xcal(wisperdata) # Returned as statsmodels results object

        ## Fitting the iso ratios requires polynomial model selection:
        """
        Using min Bayesian info criterion (BIC) to determine highest power 
        (up to 5) of each predictor var and crossterm, for the chosen isotope 
        variable. Returns a 3-tuple of ints, where each is the highest power 
        to raise the predictor vars: logq, iso, and logq*iso. iso is either 
        'dD' or 'd18O':
        """
        def polyord_minBIC(wisperdata, iso):
            # Cartesian product of all poly orders up to 5:
            nord_list = list(itertools.product(range(1,6), range(1,6), range(1,6)))
            bic_list = []
            for nord in nord_list:
                model = isoxcal_modelfit(wisperdata, iso, nord) # Statsmodels results.
                bic_list.append(model.bic) # Append this run's BIC.
            # Combo of poly orders with the minimum BIC:
            return nord_list[np.argmin(bic_list)]

            # Find polynomial orders for each iso ratio. Then re-run fit with 
            # those poly orders:
        nord_dD = polyord_minBIC(wisperdata, 'dD')
        nord_d18O = polyord_minBIC(wisperdata, 'd18O')
        model_dD = isoxcal_modelfit(wisperdata, 'dD', nord_dD)
        model_d18O = isoxcal_modelfit(wisperdata, 'd18O', nord_d18O)
        
        ## Print results:
        print("****************************************************\n"
              "****************************************************\n"
              "Cross-calibration fit parameters for ORACLES "+year+"\n"
              "****************************************************\n"
              "****************************************************")
        #print(model_q.summary())
        #print(model_dD.summary())
        #print(model_d18O.summary())
        print('\nq\n===')
        print('R2 = %f' % model_q.rsquared)

        print('\ndD\n===')
        print('nord = % s' % str(nord_dD))
        print('R2 = %f' % model_dD.rsquared)
        
        print('\nd18O\n====')
        print('nord = % s' % str(nord_d18O))
        print('R2 = %f' % model_d18O.rsquared)

    
        ## Figure of data with fit:
        draw_fitfig(year, wisperdata, model_dD.params, model_d18O.params)

        
        ## Return parameter fits:
        return {'q':model_q.params, 'dD':model_dD.params, 'd18O':model_d18O.params}


    """
    Save pandas dataframe (data) as a new excel sheet (with name 'newsheetname'), 
    to a workbook (at path='path_workbook'). If workbook doesn't exist, create it.
    """
    def add_excel_sheet(path_workbook, data, newsheetname):
        
        writer = pd.ExcelWriter(path_workbook, engine='openpyxl')
    
        if os.path.isfile(path_workbook): # Workbook exists. Add sheets.
            book = load_workbook(path_workbook)
            writer.book = book
            
            # ExcelWriter uses writer.sheets to access the sheet. If you leave it 
            # empty it will not overwrite an existing sheet with the same name.
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            
            data.to_excel(writer, newsheetname, index=False)
            writer.save()
    
        else: # Workbook doesn't exist. Create it and save dataframe.     
            data.to_excel(writer, newsheetname, index=False)
            writer.save()
    
    
    ## Get fit parameters and save to my calibration fits excel workbook:
        # Fit parameters for each year:
    fitparams_2017 = get_fits_year('2017')
    fitparams_2018 = get_fits_year('2018')
    
    """
        # For each cross-cal variable, make a single pandas df with parameters from 
        # both years:
    def make_fitparam_df(crosscalvar): # crosscalvar (str) = a dict key.
        fitpar_df = pd.DataFrame([dict(fitparams_2017[crosscalvar]), 
                                  dict(fitparams_2018[crosscalvar])], 
                                 index=['2017','2018'])
        fitpar_df['year'] = fitpar_df.index.values
        return fitpar_df
    
    params_h2o = make_fitparam_df('q')
    params_dD = make_fitparam_df('dD')
    params_d18O = make_fitparam_df('d18O')
       
        # Write to excel sheets:
    sheetnames = ['pic2pic1_xcal_h2o','pic2pic1_xcal_dD','pic2pic1_xcal_d18O']
    params_ready2write = [params_h2o, params_dD, params_d18O]

    relpath_caltable = r'../Calibration_Data/calibration_fits.xlsx' 
    for param_df, sheetname in zip(params_ready2write, sheetnames):
        add_excel_sheet(relpath_caltable, param_df, sheetname)
    """
    
##-----------------------------------------------------------------------------


if __name__=='__main__':
    get_fits()  



## Look at fit parameter covariances and correlations:
#cov = cov = model.cov_params().values # Covariance matrix.
#Dinv = np.diag(1 / np.sqrt(np.diag(cov))) 
#corr = (Dinv @ cov) @ Dinv
#print(corr)


## Publication-ready plot of 2D cross-cal maps:
##-----------------------------------------------------------------------------    
"""
    # Compute cross-calibration values on (logq, delta) grids:
logq1=np.log(100); logq2=np.log(22000); dD1=-450; dD2=0; d18O1=-55; d18O2=0
logq_grid, dD_grid = np.meshgrid((np.linspace(logq1, logq2, 200)), 
                                 np.linspace(dD1, dD2, 100))
logq_grid, d18O_grid = np.meshgrid((np.linspace(logq1, logq2, 200)), 
                                 np.linspace(d18O1, d18O2, 100))
dD_cal = xcal_dD(logq_grid, dD_grid, dict(model_dD.params))
d18O_cal = xcal_d18O(logq_grid, d18O_grid, dict(model_d18O.params))

    # Plot:
fig_pub = plt.figure()
ax1_pub = plt.subplot(1,2,1)    
ax2_pub = plt.subplot(1,2,2)

pc_dD = ax1_pub.scatter(logq_grid.flatten(), dD_grid.flatten(), 
                        c=dD_cal.flatten())
pc_d18O = ax2_pub.scatter(logq_grid.flatten(), d18O_grid.flatten(), 
                          c=d18O_cal.flatten())

fig_pub.colorbar(pc_dD, ax=ax1_pub)
fig_pub.colorbar(pc_d18O, ax=ax2_pub)

ax1_pub.set_xlim(logq1, logq2); ax1_pub.set_ylim(dD1, dD2)
ax2_pub.set_xlim(logq1, logq2); ax2_pub.set_ylim(d18O1, d18O2)

ax1_pub.set_xlabel('Pic2 log(q[ppmv])', fontsize=14)
ax1_pub.set_ylabel(r'Pic2 $\delta$D [permil]', fontsize=14)
ax2_pub.set_xlabel('Pic2 log(q[ppmv])', fontsize=14)
ax2_pub.set_ylabel(r'Pic2 $\delta^{18}$O [permil]', fontsize=14)    
"""
##----------------------------------------------------------------------------- 
